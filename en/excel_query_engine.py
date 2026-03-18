"""
Excel Query Engine - Smart search with type inference, normalization, and multi-criteria support
Production-ready implementation for 10,000+ row Excel files
"""

import re
from typing import List, Dict, Tuple, Any, Optional, Union
from dataclasses import dataclass
from enum import Enum
from datetime import datetime
import string

try:
    from excel_ai import detect_headers_upgrade
    EXCEL_AI_AVAILABLE = True
except ImportError:
    EXCEL_AI_AVAILABLE = False

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from dateutil import parser as date_parser
    DATEUTIL_AVAILABLE = True
except ImportError:
    DATEUTIL_AVAILABLE = False


# ============================================================================
# ENUMS & DATA CLASSES
# ============================================================================

class DataType(Enum):
    """Detected data types for columns"""
    NUMERIC = "numeric"
    DATE = "date"
    TEXT = "text"
    MIXED = "mixed"
    EMPTY = "empty"


class SearchMode(Enum):
    """Search matching modes"""
    EXACT = "exact"
    PARTIAL = "partial"
    BOTH = "both"


class LogicalOperator(Enum):
    """Multi-criteria logical operators"""
    AND = "AND"
    OR = "OR"


@dataclass
class ColumnProfile:
    """Profile metadata for a single column"""
    name: str
    index: int
    detected_type: DataType
    non_empty_count: int
    sample_values: List[str]
    numeric_range: Optional[Tuple[float, float]] = None
    date_range: Optional[Tuple[datetime, datetime]] = None


@dataclass
class SearchResult:
    """Result of a single search criterion"""
    criterion: Tuple[Any, str]  # (value/values, column_name)
    exact_matches: List[Dict[str, Any]]
    partial_matches: List[Dict[str, Any]]
    matched_column_name: str
    matched_column_index: int
    matched_column_type: DataType


@dataclass
class QueryResult:
    """Final result of a multi-criteria query"""
    mode: LogicalOperator
    total_matches: int
    matched_rows: List[Dict[str, Any]]
    individual_results: List[SearchResult]
    query_summary: str


# ============================================================================
# DATA NORMALIZER
# ============================================================================

class DataNormalizer:
    """Normalize and clean Excel data"""
    
    DATE_FORMATS = [
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%m-%d-%Y",
        "%Y/%m/%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d.%m.%Y",
        "%Y.%m.%d",
        "%B %d, %Y",
        "%b %d, %Y",
        "%d %B %Y",
        "%d %b %Y",
    ]
    
    @staticmethod
    def normalize_cell(value: Any) -> str:
        """
        Normalize a single cell value:
        - Convert to string
        - Strip whitespace
        - Handle None/null
        """
        if value is None or value == "":
            return ""
        
        # Convert to string and strip
        normalized = str(value).strip()
        return normalized
    
    @staticmethod
    def try_parse_date(value: str) -> Optional[datetime]:
        """Try to parse string as date"""
        if not DATEUTIL_AVAILABLE:
            return None
        
        value = value.strip()
        if not value:
            return None
        
        try:
            # Try dateutil parser first (handles many formats)
            parsed = date_parser.parse(value, dayfirst=False, fuzzy=False)
            return parsed
        except (ValueError, TypeError, OverflowError):
            pass
        
        # Try manual formats
        for fmt in DataNormalizer.DATE_FORMATS:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
        
        return None
    
    @staticmethod
    def try_parse_number(value: str) -> Optional[float]:
        """Try to parse string as number"""
        value = value.strip()
        if not value:
            return None
        
        try:
            # Try int first
            if '.' not in value and ',' not in value:
                return float(int(value))
            # Try float
            return float(value.replace(',', '.'))
        except (ValueError, TypeError):
            return None
    
    @staticmethod
    def infer_type(value: str) -> DataType:
        """Infer data type of normalized value"""
        if not value:
            return DataType.EMPTY
        
        # Try numeric
        if DataNormalizer.try_parse_number(value) is not None:
            return DataType.NUMERIC
        
        # Try date
        if DataNormalizer.try_parse_date(value) is not None:
            return DataType.DATE
        
        # Default to text
        return DataType.TEXT


# ============================================================================
# COLUMN PROFILER
# ============================================================================

class ColumnProfiler:
    """Profile columns for type detection and statistics"""
    
    @staticmethod
    def profile_column(column_data: List[str], column_name: str, column_index: int) -> ColumnProfile:
        """
        Profile a single column:
        - Detect dominant type
        - Gather statistics
        - Store sample values
        """
        type_counts = {dt: 0 for dt in DataType}
        numeric_values = []
        date_values = []
        non_empty = 0
        
        for value in column_data:
            if value:
                non_empty += 1
                inferred_type = DataNormalizer.infer_type(value)
                type_counts[inferred_type] += 1
                
                # Collect numeric values
                if inferred_type == DataType.NUMERIC:
                    num = DataNormalizer.try_parse_number(value)
                    if num is not None:
                        numeric_values.append(num)
                
                # Collect date values
                if inferred_type == DataType.DATE:
                    date = DataNormalizer.try_parse_date(value)
                    if date is not None:
                        date_values.append(date)
        
        # Determine dominant type
        relevant_types = [dt for dt in [DataType.NUMERIC, DataType.DATE, DataType.TEXT] 
                         if type_counts[dt] > 0]
        
        if len(relevant_types) == 1:
            dominant_type = relevant_types[0]
        elif len(relevant_types) > 1:
            dominant_type = DataType.MIXED
        else:
            dominant_type = DataType.EMPTY
        
        # Calculate ranges
        numeric_range = None
        if numeric_values:
            numeric_range = (min(numeric_values), max(numeric_values))
        
        date_range = None
        if date_values:
            date_range = (min(date_values), max(date_values))
        
        # Sample values
        sample_values = [v for v in column_data if v][:5]
        
        return ColumnProfile(
            name=column_name,
            index=column_index,
            detected_type=dominant_type,
            non_empty_count=non_empty,
            sample_values=sample_values,
            numeric_range=numeric_range,
            date_range=date_range
        )
    
    @staticmethod
    def profile_sheet(rows: List[List[str]], headers: List[str]) -> List[ColumnProfile]:
        """Profile all columns in sheet"""
        profiles = []
        
        # Extract column data
        num_cols = len(headers)
        for col_idx in range(num_cols):
            column_data = [row[col_idx] if col_idx < len(row) else "" for row in rows]
            profile = ColumnProfiler.profile_column(column_data, headers[col_idx], col_idx)
            profiles.append(profile)
        
        return profiles


# ============================================================================
# COLUMN MATCHER (Fuzzy Header Matching)
# ============================================================================

class ColumnMatcher:
    """Fuzzy match column hints to actual headers"""
    
    @staticmethod
    def levenshtein_distance(s1: str, s2: str) -> int:
        """Calculate Levenshtein distance between two strings"""
        s1 = s1.lower().strip()
        s2 = s2.lower().strip()
        
        if len(s1) < len(s2):
            s1, s2 = s2, s1
        
        if len(s2) == 0:
            return len(s1)
        
        previous_row = range(len(s2) + 1)
        for i, c1 in enumerate(s1):
            current_row = [i + 1]
            for j, c2 in enumerate(s2):
                insertions = previous_row[j + 1] + 1
                deletions = current_row[j] + 1
                substitutions = previous_row[j] + (c1 != c2)
                current_row.append(min(insertions, deletions, substitutions))
            previous_row = current_row
        
        return previous_row[-1]
    
    @staticmethod
    def find_column(hint: str, headers: List[str], max_distance: int = 2) -> Tuple[Optional[str], Optional[int], float]:
        """
        Find column matching hint with fuzzy matching
        
        Returns: (matched_column_name, column_index, confidence_score)
        - confidence_score: 1.0 = exact match, 0.0 = no match
        """
        if not hint or not headers:
            return None, None, 0.0
        
        hint_lower = hint.lower().strip()
        
        # First, try exact match (case-insensitive)
        for idx, header in enumerate(headers):
            if header.lower().strip() == hint_lower:
                return header, idx, 1.0
        
        # Then try fuzzy match
        matches = []
        for idx, header in enumerate(headers):
            distance = ColumnMatcher.levenshtein_distance(hint, header)
            if distance <= max_distance:
                # Confidence: inverse of distance
                confidence = 1.0 - (distance / max(len(hint), len(header)))
                matches.append((header, idx, confidence, distance))
        
        if matches:
            # Sort by confidence (descending)
            matches.sort(key=lambda x: (-x[2], x[3]))
            best_match = matches[0]
            
            # Warn if multiple matches with similar confidence
            if len(matches) > 1 and abs(matches[0][2] - matches[1][2]) < 0.1:
                print(f"⚠️  WARNING: Ambiguous column hint '{hint}'")
                print(f"   Top matches: {matches[0][0]} ({matches[0][2]:.1%}), {matches[1][0]} ({matches[1][2]:.1%})")
                print(f"   Using: {matches[0][0]}")
            
            return best_match[0], best_match[1], best_match[2]
        
        return None, None, 0.0


# ============================================================================
# SEARCH ENGINE
# ============================================================================

class SearchEngine:
    """Core search engine with type inference and matching strategies"""
    
    @staticmethod
    def infer_search_term_type(term: Any) -> DataType:
        """Infer type of search term"""
        if isinstance(term, (int, float)):
            return DataType.NUMERIC
        
        term_str = str(term).strip()
        
        # Try numeric
        if DataNormalizer.try_parse_number(term_str) is not None:
            return DataType.NUMERIC
        
        # Try date
        if DataNormalizer.try_parse_date(term_str) is not None:
            return DataType.DATE
        
        return DataType.TEXT
    
    @staticmethod
    def should_search_column(column_profile: ColumnProfile, search_type: DataType) -> bool:
        """Determine if column should be searched based on type compatibility"""
        col_type = column_profile.detected_type
        
        # Always search mixed columns
        if col_type == DataType.MIXED:
            return True
        
        # Don't search empty columns
        if col_type == DataType.EMPTY:
            return False
        
        # Match types
        if search_type == col_type:
            return True
        
        # Text search can match any non-empty column
        if search_type == DataType.TEXT:
            return True
        
        return False
    
    @staticmethod
    def normalize_for_comparison(value: str, value_type: DataType) -> str:
        """Normalize value for comparison"""
        value = value.strip()
        
        if value_type == DataType.NUMERIC:
            # Normalize number: remove leading zeros (except single 0)
            try:
                num = DataNormalizer.try_parse_number(value)
                if num is not None:
                    if num == int(num):
                        return str(int(num))
                    else:
                        return str(num)
            except:
                pass
        
        if value_type == DataType.DATE:
            # Normalize date to ISO format
            date = DataNormalizer.try_parse_date(value)
            if date:
                return date.strftime("%Y-%m-%d")
        
        return value
    
    @staticmethod
    def exact_match(search_term: Any, cell_value: str, search_type: DataType) -> bool:
        """Check for exact match"""
        search_str = str(search_term).strip()
        cell_normalized = SearchEngine.normalize_for_comparison(cell_value, search_type)
        search_normalized = SearchEngine.normalize_for_comparison(search_str, search_type)
        
        return cell_normalized == search_normalized
    
    @staticmethod
    def partial_match(search_term: Any, cell_value: str, search_type: DataType) -> bool:
        """Check for partial match (case-insensitive substring)"""
        search_str = str(search_term).strip().lower()
        cell_lower = cell_value.strip().lower()
        
        return search_str in cell_lower and search_str != cell_lower
    
    @staticmethod
    def search_column(
        search_term: Any,
        rows: List[Dict[str, str]],
        column_name: str,
        search_type: DataType
    ) -> Tuple[List[Dict[str, str]], List[Dict[str, str]]]:
        """
        Search column for term
        
        Returns: (exact_matches, partial_matches)
        """
        exact_matches = []
        partial_matches = []
        
        for row in rows:
            if column_name not in row:
                continue
            
            cell_value = row[column_name]
            
            # Check exact match
            if SearchEngine.exact_match(search_term, cell_value, search_type):
                exact_matches.append(row)
            # Check partial match
            elif SearchEngine.partial_match(search_term, cell_value, search_type):
                partial_matches.append(row)
        
        return exact_matches, partial_matches


# ============================================================================
# QUERY BUILDER (Multi-Criteria Orchestration)
# ============================================================================

class QueryBuilder:
    """Build and execute multi-criteria queries with AND/OR logic"""
    
    def __init__(self, rows: List[Dict[str, str]], headers: List[str], profiles: List[ColumnProfile]):
        self.rows = rows
        self.headers = headers
        self.profiles = profiles
    
    def search_single_criterion(
        self,
        search_values: Union[Any, List[Any]],
        column_hint: str
    ) -> SearchResult:
        """
        Search for single criterion (may have multiple values)
        
        Args:
            search_values: Single value or list of values to search for
            column_hint: Column name or hint
        
        Returns:
            SearchResult with exact and partial matches
        """
        # Normalize search values to list
        if not isinstance(search_values, list):
            search_values = [search_values]
        
        # Find target column
        column_name, column_idx, confidence = ColumnMatcher.find_column(column_hint, self.headers)
        
        if column_name is None:
            raise ValueError(f"❌ Column hint '{column_hint}' did not match any column. Available: {self.headers}")
        
        if confidence < 0.5:
            print(f"⚠️  Low confidence match for '{column_hint}' -> '{column_name}' ({confidence:.1%})")
        
        # Get column profile
        column_profile = self.profiles[column_idx]
        
        # Infer search type from first value
        search_type = SearchEngine.infer_search_term_type(search_values[0])
        
        # Search for all values (OR logic within same criterion)
        all_exact = []
        all_partial = []
        
        for search_term in search_values:
            exact, partial = SearchEngine.search_column(
                search_term,
                self.rows,
                column_name,
                search_type
            )
            all_exact.extend(exact)
            all_partial.extend(partial)
        
        # Remove duplicates while preserving order
        seen_exact = set()
        unique_exact = []
        for row in all_exact:
            row_tuple = tuple(row.items())
            if row_tuple not in seen_exact:
                seen_exact.add(row_tuple)
                unique_exact.append(row)
        
        seen_partial = set()
        unique_partial = []
        for row in all_partial:
            row_tuple = tuple(row.items())
            if row_tuple not in seen_exact and row_tuple not in seen_partial:
                seen_partial.add(row_tuple)
                unique_partial.append(row)
        
        return SearchResult(
            criterion=(search_values, column_hint),
            exact_matches=unique_exact,
            partial_matches=unique_partial,
            matched_column_name=column_name,
            matched_column_index=column_idx,
            matched_column_type=column_profile.detected_type
        )
    
    def query(
        self,
        criteria: List[Tuple[Union[Any, List[Any]], str]],
        mode: LogicalOperator = LogicalOperator.AND,
        include_partial: bool = False
    ) -> QueryResult:
        """
        Execute multi-criteria query
        
        Args:
            criteria: List of (search_values, column_hint) tuples
            mode: AND or OR
            include_partial: Whether to include partial matches
        
        Returns:
            QueryResult with matched rows
        """
        individual_results = []
        
        # Search each criterion
        for search_values, column_hint in criteria:
            result = self.search_single_criterion(search_values, column_hint)
            individual_results.append(result)
        
        # Combine results based on mode
        if mode == LogicalOperator.AND:
            matched_rows = self._apply_and_logic(individual_results, include_partial)
        else:  # OR
            matched_rows = self._apply_or_logic(individual_results, include_partial)
        
        # Create summary
        summary = self._create_summary(criteria, mode, individual_results, len(matched_rows))
        
        return QueryResult(
            mode=mode,
            total_matches=len(matched_rows),
            matched_rows=matched_rows,
            individual_results=individual_results,
            query_summary=summary
        )
    
    @staticmethod
    def _apply_and_logic(results: List[SearchResult], include_partial: bool) -> List[Dict[str, str]]:
        """AND mode: return rows matching ALL criteria"""
        if not results:
            return []
        
        # Start with rows from first criterion
        first_result = results[0]
        candidate_rows = set()
        for row in first_result.exact_matches:
            candidate_rows.add(tuple(row.items()))
        if include_partial:
            for row in first_result.partial_matches:
                candidate_rows.add(tuple(row.items()))
        
        # Intersect with subsequent criteria
        for result in results[1:]:
            criterion_rows = set()
            for row in result.exact_matches:
                criterion_rows.add(tuple(row.items()))
            if include_partial:
                for row in result.partial_matches:
                    criterion_rows.add(tuple(row.items()))
            
            candidate_rows = candidate_rows.intersection(criterion_rows)
        
        # Convert back to dicts
        return [dict(row) for row in candidate_rows]
    
    @staticmethod
    def _apply_or_logic(results: List[SearchResult], include_partial: bool) -> List[Dict[str, str]]:
        """OR mode: return rows matching ANY criterion"""
        candidate_rows = {}  # Use dict to maintain order and avoid duplicates
        
        for result in results:
            for row in result.exact_matches:
                row_tuple = tuple(row.items())
                if row_tuple not in candidate_rows:
                    candidate_rows[row_tuple] = dict(row)
            
            if include_partial:
                for row in result.partial_matches:
                    row_tuple = tuple(row.items())
                    if row_tuple not in candidate_rows:
                        candidate_rows[row_tuple] = dict(row)
        
        return list(candidate_rows.values())
    
    @staticmethod
    def _create_summary(criteria, mode, results, total_matches) -> str:
        """Create human-readable query summary"""
        criteria_strs = []
        for i, result in enumerate(results):
            search_vals, col_hint = criteria[i]
            if isinstance(search_vals, list):
                vals_str = " OR ".join(str(v) for v in search_vals)
            else:
                vals_str = str(search_vals)
            
            exact_count = len(result.exact_matches)
            partial_count = len(result.partial_matches)
            
            criteria_strs.append(
                f"  • [{result.matched_column_name}] = {vals_str} "
                f"({exact_count} exact, {partial_count} partial)"
            )
        
        return (
            f"Query Mode: {mode.value}\n"
            f"Criteria:\n" + "\n".join(criteria_strs) + f"\n"
            f"Total Matches: {total_matches}"
        )


# ============================================================================
# EXCEL READER
# ============================================================================

class ExcelReader:
    """Read and normalize Excel files"""
    
    @staticmethod
    def read_file(filepath: str) -> Tuple[List[str], List[List[str]]]:
        """
        Read Excel file and return normalized data
        
        Returns: (headers, rows)
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Read all data
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append(row)
        
        if not all_rows:
            raise ValueError("Excel file is empty")
        
        # Extract headers
        header_row_idx = 0

        if EXCEL_AI_AVAILABLE:
            # Detect headers using upgraded model
            header_info = detect_headers_upgrade(filepath)

            sheet_name = list(header_info.keys())[0]
            info = header_info[sheet_name]

            header_rows = info["header_rows"]
            confidence = info["confidence"]
            reconstructed_columns = info["columns"]

            print(f"🧠 Header rows detected: {header_rows} (confidence={confidence:.2f})")

            # Convert to 0-based index
            header_row_idx = header_rows[-1] - 1

            # ----------------------------------------
            # CASE 1: reconstructed header exists
            # ----------------------------------------
            if reconstructed_columns is not None:
                headers = [
                    DataNormalizer.normalize_cell(h)
                    for h in reconstructed_columns
                ]

            # ----------------------------------------
            # CASE 2: fallback to detected row
            # ----------------------------------------
            else:
                headers = [
                    DataNormalizer.normalize_cell(h)
                    for h in all_rows[header_row_idx]
                ]

            # ----------------------------------------
            # SAFETY: low confidence fallback
            # ----------------------------------------
            if confidence < 0.5:
                print("⚠️ Low confidence → fallback to row 1")
                header_row_idx = 0
                headers = [
                    DataNormalizer.normalize_cell(h)
                    for h in all_rows[0]
                ]
        else:
            # Fallback: use first row as headers
            print("ℹ️ excel_ai not available — using first row as headers")
            header_row_idx = 0
            headers = [
                DataNormalizer.normalize_cell(h)
                for h in all_rows[0]
            ]
        
        # Normalize data rows and convert to dicts
        normalized_rows = []
        for row_data in all_rows[header_row_idx + 1:]:
            normalized_row = {}
            for col_idx, header in enumerate(headers):
                if col_idx < len(row_data):
                    normalized_row[header] = DataNormalizer.normalize_cell(row_data[col_idx])
                else:
                    normalized_row[header] = ""
            normalized_rows.append(normalized_row)
        
        return headers, normalized_rows


# ============================================================================
# EXCEL EXPORTER
# ============================================================================

class ExcelExporter:
    """Export query results to Excel"""
    
    @staticmethod
    def export_results(
        filepath: str,
        headers: List[str],
        rows: List[Dict[str, str]],
        query_summary: str = ""
    ):
        """Export matched rows to new Excel file"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        
        # Add headers
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Add data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Add metadata sheet
        if query_summary:
            metadata_ws = wb.create_sheet("Query Metadata")
            metadata_ws['A1'] = "Query Summary"
            metadata_ws['A2'] = query_summary
            metadata_ws['A2'].alignment = metadata_ws['A2'].alignment.copy()
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
        print(f"✅ Results exported to: {filepath}")


# ============================================================================
# MAIN QUERY ENGINE
# ============================================================================

class ExcelQueryEngine:
    """Main interface for Excel query operations"""
    
    def __init__(self, excel_filepath: str):
        """Initialize engine with Excel file"""
        print(f"📂 Loading Excel file: {excel_filepath}")
        self.filepath = excel_filepath
        
        # Read file
        self.headers, self.rows = ExcelReader.read_file(excel_filepath)
        print(f"✅ Loaded {len(self.rows)} rows, {len(self.headers)} columns")
        
        # Profile columns
        print("📊 Profiling columns...")
        self.profiles = ColumnProfiler.profile_sheet([list(r.values()) for r in self.rows], self.headers)
        
        for profile in self.profiles:
            print(f"   • {profile.name:20s} | Type: {profile.detected_type.value:7s} | "
                  f"Non-empty: {profile.non_empty_count}/{len(self.rows)}")
        
        self.query_builder = QueryBuilder(self.rows, self.headers, self.profiles)
    
    def search(
        self,
        criteria: List[Tuple[Union[Any, List[Any]], str]],
        mode: LogicalOperator = LogicalOperator.AND,
        include_partial: bool = False,
        export_filepath: Optional[str] = None
    ) -> QueryResult:
        """
        Execute search query
        
        Args:
            criteria: List of (search_values, column_hint) tuples
            mode: AND or OR
            include_partial: Include partial matches
            export_filepath: Optional Excel file to export results
        
        Returns:
            QueryResult with matched rows
        """
        print(f"\n🔍 Executing query ({mode.value} mode)...")
        result = self.query_builder.query(criteria, mode, include_partial)
        
        print(f"\n{result.query_summary}")
        print(f"✅ Found {result.total_matches} matching rows")
        
        # Export if requested
        if export_filepath:
            ExcelExporter.export_results(
                export_filepath,
                self.headers,
                result.matched_rows,
                result.query_summary
            )
        
        return result
    
    def get_column_info(self, column_hint: str) -> ColumnProfile:
        """Get detailed info about a column"""
        col_name, col_idx, confidence = ColumnMatcher.find_column(column_hint, self.headers)
        if col_idx is None:
            raise ValueError(f"Column '{column_hint}' not found")
        return self.profiles[col_idx]


# ============================================================================
# EXAMPLE USAGE
# ============================================================================

if __name__ == "__main__":
    print("=" * 80)
    print("EXCEL QUERY ENGINE - DEMO")
    print("=" * 80)
    
    # Example usage (requires test_data.xlsx)
    try:
        engine = ExcelQueryEngine("test_data.xlsx")
        
        # Example 1: Simple search
        print("\n" + "=" * 80)
        print("EXAMPLE 1: Search for 'John' in Name column")
        print("=" * 80)
        result1 = engine.search(
            criteria=[("John", "Name")],
            include_partial=True,
            export_filepath="example1_results.xlsx"
        )
        
        # Example 2: AND query
        print("\n" + "=" * 80)
        print("EXAMPLE 2: AND query - Name='John' AND Age=30")
        print("=" * 80)
        result2 = engine.search(
            criteria=[("John", "Name"), (30, "Age")],
            mode=LogicalOperator.AND,
            include_partial=False,
            export_filepath="example2_results.xlsx"
        )
        
        # Example 3: OR query
        print("\n" + "=" * 80)
        print("EXAMPLE 3: OR query - Name='John' OR Name='Jane'")
        print("=" * 80)
        result3 = engine.search(
            criteria=[
                (["John", "Jane"], "Name")
            ],
            mode=LogicalOperator.OR,
            include_partial=False,
            export_filepath="example3_results.xlsx"
        )
        
        # Example 4: Multiple criteria with multiple values
        print("\n" + "=" * 80)
        print("EXAMPLE 4: Complex - (Name='John' OR Name='Jane') AND Salary >= 50000")
        print("=" * 80)
        result4 = engine.search(
            criteria=[
                (["John", "Jane"], "Name"),
                ([50000, 60000, 70000], "Salary")
            ],
            mode=LogicalOperator.AND,
            include_partial=False,
            export_filepath="example4_results.xlsx"
        )
        
    except FileNotFoundError:
        print("⚠️  test_data.xlsx not found. Create it first or provide a valid Excel file.")
        print("\nUsage example:")
        print("  engine = ExcelQueryEngine('your_file.xlsx')")
        print("  result = engine.search(")
        print("      criteria=[('search_value', 'Column Name')],")
        print("      mode=LogicalOperator.AND,")
        print("      export_filepath='results.xlsx'")
        print("  )")